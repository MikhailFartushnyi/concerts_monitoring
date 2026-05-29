"""
Работа с xlsx.
COLUMNS в config.py — словарь {ключ: заголовок}.
"""

import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_FILE, COLUMNS, DATA_DIR

KEYS = list(COLUMNS.keys())
HEADERS = list(COLUMNS.values())

FILL_NEW = PatternFill("solid", fgColor="E8F5E9")
FILL_DUPE = PatternFill("solid", fgColor="FFF9C4")
FILL_HEADER = PatternFill("solid", fgColor="1565C0")
FONT_HEADER = Font(color="FFFFFF", bold=True)
FONT_LINK = Font(color="1976D2", underline="single")


def load_existing_ids() -> set[str]:
    if not os.path.exists(OUTPUT_FILE):
        return set()
    try:
        df = pd.read_excel(OUTPUT_FILE, usecols=["ID"], dtype=str)
        return set(df["ID"].dropna().tolist())
    except Exception:
        return set()


def save_new_records(new_records: list[dict]) -> int:
    os.makedirs(DATA_DIR, exist_ok=True)

    if not new_records:
        return 0

    # Строим DataFrame по ключам, переименовываем в заголовки
    df_new = pd.DataFrame(new_records, columns=KEYS)
    df_new.columns = HEADERS

    if os.path.exists(OUTPUT_FILE):
        df_exist = pd.read_excel(OUTPUT_FILE, dtype=str)
        existing_rows = len(df_exist)
        df_all = pd.concat([df_exist, df_new], ignore_index=True)
    else:
        existing_rows = 0
        df_all = df_new

    df_all.to_excel(OUTPUT_FILE, index=False, engine="openpyxl")
    _apply_styles(existing_rows, len(df_new))
    return len(df_new)


def _apply_styles(existing_rows: int, new_rows: int):
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Шапка
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30

    # Индексы нужных столбцов по заголовку
    header_to_col = {cell.value: cell.column for cell in ws[1]}
    col_url = header_to_col.get("Ссылка")
    col_photo = header_to_col.get("Фото")
    col_dupe = header_to_col.get("Дубли")

    # Ширины
    col_widths = {
        "ID": 14, "Ссылка": 45, "Источник": 14,
        "Заголовок": 45, "Название": 35, "Описание": 50,
        "Площадка": 25, "Адрес": 30, "Дата": 18, "Возраст": 8,
        "Цена старая": 14, "Цена со скидкой": 16, "Скидка": 8,
        "Промокод": 16, "Фото": 50, "Найден": 12, "Дубли": 30,
    }
    for h, w in col_widths.items():
        col_idx = header_to_col.get(h)
        if col_idx:
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Данные
    for row_idx in range(2, existing_rows + new_rows + 2):
        is_new = row_idx > existing_rows + 1
        has_dupe = False
        if col_dupe:
            v = ws.cell(row=row_idx, column=col_dupe).value
            has_dupe = bool(v and str(v).strip())

        for cell in ws[row_idx]:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if has_dupe:
                cell.fill = FILL_DUPE
            elif is_new:
                cell.fill = FILL_NEW

        if col_url:
            c = ws.cell(row=row_idx, column=col_url)
            if c.value and str(c.value).startswith("http"):
                c.hyperlink = str(c.value)
                c.font = FONT_LINK

        if col_photo:
            c = ws.cell(row=row_idx, column=col_photo)
            if c.value and str(c.value).startswith("http"):
                c.hyperlink = str(c.value)
                c.font = FONT_LINK

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(OUTPUT_FILE)
